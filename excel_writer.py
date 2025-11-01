# excel_writer.py
from __future__ import annotations
from pathlib import Path
from typing import Dict, Optional, Iterable, Tuple, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from copy import copy as _copy


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    """Mapeia 'Texto do Cabeçalho' -> índice da coluna, lendo a linha de cabeçalho."""
    row_vals = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    header_map: Dict[str, int] = {}
    for idx, val in enumerate(row_vals, start=1):
        if val is None:
            continue
        header_map[str(val).strip()] = idx
    return header_map


def _used_data_region(ws, start_row: int, max_col: int) -> Tuple[int, int]:
    """
    Determina a região de dados 'usada' a partir de start_row:
      retorna (first_data_row, last_data_row_existente)
    Se não houver dados, last_data_row = start_row - 1.
    """
    last = start_row - 1
    for r in range(ws.max_row, start_row - 1, -1):
        row_has_value = False
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                row_has_value = True
                break
        if row_has_value:
            last = r
            break
    return start_row, last


def _clear_values_preserving_styles(ws, start_row: int, max_col: int):
    """
    Limpa SOMENTE os valores das células da área de dados usada,
    preservando bordas/preenchimentos/alinhamentos/mesclas/validações.
    """
    first, last = _used_data_region(ws, start_row, max_col)
    if last < first:
        return  # nada a limpar
    for r in range(first, last + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def _copy_cell_style(src: Cell, dst: Cell):
    """
    Copia estilo de uma célula modelo para outra.
    Usa copy() para evitar 'StyleProxy' não-hashable em alguns ambientes.
    Tem fallback para named style se existir.
    """
    try:
        # Objetos de estilo devem ser copiados
        if src.fill is not None:
            dst.fill = _copy(src.fill)
        if src.border is not None:
            dst.border = _copy(src.border)
        if src.font is not None:
            dst.font = _copy(src.font)
        if src.alignment is not None:
            dst.alignment = _copy(src.alignment)

        # number_format é string; pode atribuir direto
        if src.number_format:
            dst.number_format = src.number_format

        if src.protection is not None:
            dst.protection = _copy(src.protection)

    except TypeError:
        # Fallback: se houver named style, aplica
        if getattr(src, "style", None):
            dst.style = src.style
        # Caso não haja named style, o destino fica com o estilo atual (já preservado)


def _ensure_row_styles(ws, template_row: int, target_row: int, columns: List[int]):
    """
    Replica estilos da 'template_row' apenas nas colunas especificadas.
    """
    for c in columns:
        _copy_cell_style(ws.cell(row=template_row, column=c), ws.cell(row=target_row, column=c))


def write_df_to_existing_template(
    xlsx_path: str | Path,
    df: pd.DataFrame,
    *,
    sheet_name: str = "FROTA-Layout_excel_Geral",
    header_row: int = 4,           # << ajuste conforme seu layout
    data_start_row: int = 6,
    column_map: Optional[Dict[str, str]] = None,
    strict: bool = False,
    template_row_for_style: Optional[int] = None,  # se None, usa data_start_row como modelo
) -> None:
    """
    Escreve df no arquivo Excel EXISTENTE, na aba 'sheet_name', preservando layout.
    - Não cria aba nova.
    - Limpa apenas a área de dados usada (valores), mantendo estilos.
    - Casa colunas do df com os cabeçalhos da planilha (ou usa column_map).
    - Replica o estilo da linha-modelo nas NOVAS linhas antes de escrever valores.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {xlsx_path}")

    # Preserva macros/tema quando .xlsm
    keep_vba = xlsx_path.suffix.lower() == ".xlsm"
    wb = load_workbook(xlsx_path, keep_vba=keep_vba, data_only=False)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada em {xlsx_path.name}.")
    ws = wb[sheet_name]

    header_map = _build_header_map(ws, header_row)
    max_col = ws.max_column

    # 1) Monta o mapeamento de escrita: df_col -> col_idx
    to_write: List[tuple[str, int]] = []
    missing: List[str] = []

    for col in df.columns:
        target = column_map.get(col, col) if column_map else col
        if target in header_map:
            to_write.append((col, header_map[target]))
        else:
            missing.append(col)

    if missing and strict:
        raise KeyError(
            "Colunas do DataFrame sem correspondência no cabeçalho da planilha: "
            + ", ".join(missing)
        )

    # Apenas as colunas que serão escritas precisam replicar estilo
    cols_for_style = sorted({idx for _, idx in to_write}) or list(range(1, max_col + 1))

    # 2) Limpa os VALORES da área de dados usada
    _clear_values_preserving_styles(ws, start_row=data_start_row, max_col=max_col)

    # 3) Replica estilo da linha-modelo para as linhas a serem escritas
    model_row = template_row_for_style or data_start_row
    total_rows = len(df.index)
    if total_rows > 0:
        for i in range(total_rows):
            excel_row = data_start_row + i
            _ensure_row_styles(ws, template_row=model_row, target_row=excel_row, columns=cols_for_style)

    # 4) Escreve valores nas colunas mapeadas
    for r_offset, (_, row) in enumerate(df.iterrows(), start=0):
        excel_row = data_start_row + r_offset
        for df_col, xls_col_idx in to_write:
            ws.cell(row=excel_row, column=xls_col_idx).value = row[df_col]

    # 5) Salva sem alterar layout
    wb.save(xlsx_path)
